"""
Lexus Excel Updater
Reads lexus_combined.json and updates lexus_cars.xlsx with price history tracking.
"""

import json, os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(DATA_DIR, "lexus_cars.xlsx")
JSON_PATH  = os.path.join(DATA_DIR, "lexus_combined.json")

COLUMNS = [
    ("منبع",               "source"),
    ("مدل",                "model"),
    ("واریانت",             "variant"),
    ("عنوان آگهی",          "title"),
    ("سال",                "year"),
    ("اولین ثبت",           "first_registration"),
    ("کیلومتر",             "mileage"),
    ("رنگ",                "color"),
    ("سوخت",               "fuel"),
    ("گیربکس",              "transmission"),
    ("قیمت (€)",            "price_eur"),
    ("قیمت قبلی (€)",       "prev_price_eur"),
    ("تغییر قیمت (€)",      "price_change"),
    ("ارزیابی قیمت",        "price_label"),
    ("سانروف/پانوراما",     "sunroof"),
    ("بدون تصادف",          "no_accident"),
    ("تعداد مالکان قبلی",   "previous_owners"),
    ("شهر",                "city"),
    ("کشور",               "country"),
    ("توضیحات (آلمانی)",    "description_de"),
    ("توضیحات (فارسی)",     "description_fa"),
    ("لینک",               "link"),
    ("اولین مشاهده",        "first_seen"),
    ("آخرین مشاهده",        "last_seen"),
    ("وضعیت",              "status"),
]

STATUS_COLORS = {
    "جدید":         "C6EFCE",
    "تغییر قیمت":   "FFEB9C",
    "بدون تغییر":   "FFFFFF",
    "حذف شده":      "FFC7CE",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
CELL_FONT   = Font(name="Calibri", size=9)
WRAP_ALIGN  = Alignment(wrap_text=True, vertical="top")
THIN        = Side(style="thin", color="CCCCCC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_WIDTHS  = {
    "منبع": 12, "مدل": 10, "واریانت": 12, "عنوان آگهی": 35,
    "سال": 6, "اولین ثبت": 10, "کیلومتر": 12, "رنگ": 10,
    "سوخت": 16, "گیربکس": 12, "قیمت (€)": 10, "قیمت قبلی (€)": 12,
    "تغییر قیمت (€)": 14, "ارزیابی قیمت": 14, "سانروف/پانوراما": 16,
    "بدون تصادف": 12, "تعداد مالکان قبلی": 16, "شهر": 14, "کشور": 7,
    "توضیحات (آلمانی)": 40, "توضیحات (فارسی)": 40,
    "لینک": 20, "اولین مشاهده": 12, "آخرین مشاهده": 12, "وضعیت": 12,
}


def col_for(field):
    for i, (_, f) in enumerate(COLUMNS, 1):
        if f == field:
            return i
    return None


def load_or_create_wb():
    if os.path.exists(EXCEL_PATH):
        return load_workbook(EXCEL_PATH), True
    wb = Workbook()
    ws = wb.active
    ws.title = "خودروها"
    for col, (header, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    return wb, False


def get_existing_rows(ws):
    link_col = col_for("link")
    price_col = col_for("price_eur")
    status_col = col_for("status")
    rows = {}
    if not link_col:
        return rows, link_col, price_col, status_col
    for row in range(2, ws.max_row + 1):
        link = ws.cell(row=row, column=link_col).value
        if link:
            rows[link] = {
                "row": row,
                "price": ws.cell(row=row, column=price_col).value if price_col else None,
                "status": ws.cell(row=row, column=status_col).value if status_col else "",
            }
    return rows, link_col, price_col, status_col


def write_row(ws, row_num, car, status, prev_price=None):
    today = datetime.now().strftime("%Y-%m-%d")
    car["status"] = status
    car["price_change"] = (car["price_eur"] - prev_price) if prev_price and car.get("price_eur") else None
    car["prev_price_eur"] = prev_price
    if status == "جدید":
        car["first_seen"] = today
    car["last_seen"] = today

    fill = PatternFill("solid", fgColor=STATUS_COLORS.get(status, "FFFFFF"))
    for col, (_, field) in enumerate(COLUMNS, 1):
        val = car.get(field, "") or ""
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = CELL_FONT
        cell.alignment = WRAP_ALIGN
        cell.border = BORDER
        cell.fill = fill
    ws.row_dimensions[row_num].height = 40


def run():
    with open(JSON_PATH, encoding="utf-8") as f:
        data = json.load(f)
    new_listings = data.get("listings", [])
    print(f"Loaded {len(new_listings)} listings")

    wb, existed = load_or_create_wb()
    ws = wb.active
    today = datetime.now().strftime("%Y-%m-%d")

    if existed:
        existing, link_col, price_col, status_col = get_existing_rows(ws)
        new_links = {c["link"] for c in new_listings if c.get("link")}

        # Mark removed
        for link, info in existing.items():
            if link not in new_links and info["status"] != "حذف شده":
                ws.cell(row=info["row"], column=status_col).value = "حذف شده"
                ws.cell(row=info["row"], column=status_col).fill = PatternFill("solid", fgColor=STATUS_COLORS["حذف شده"])
                lsc = col_for("last_seen")
                if lsc:
                    ws.cell(row=info["row"], column=lsc).value = today

        next_row = ws.max_row + 1
        for car in new_listings:
            link = car.get("link", "")
            if link and link in existing:
                prev = existing[link]
                old_price = prev["price"]
                new_price = car.get("price_eur")
                if old_price and new_price and int(old_price) != int(new_price):
                    write_row(ws, prev["row"], car, "تغییر قیمت", prev_price=int(old_price))
                else:
                    lsc = col_for("last_seen")
                    if lsc:
                        ws.cell(row=prev["row"], column=lsc).value = today
                    sc = col_for("status")
                    if sc:
                        ws.cell(row=prev["row"], column=sc).value = "بدون تغییر"
            else:
                write_row(ws, next_row, car, "جدید")
                next_row += 1
    else:
        for i, car in enumerate(new_listings, 2):
            write_row(ws, i, car, "جدید")

    # Column widths
    for col, (header, _) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS.get(header, 14)

    wb.save(EXCEL_PATH)
    print(f"Excel saved: {EXCEL_PATH}")


if __name__ == "__main__":
    run()
